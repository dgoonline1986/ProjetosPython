from openpyxl import Workbook
from openpyxl import load_workbook


wb=Workbook()

ws=wb.active
#nome da planilha 
ws.title

#mudando nome da planilha 
ws.title="Minha Planilha"

ws.title
#criando uma nova sheet - nome e posição
ws1=wb.create_chartsheet('teste',0)

#verificando nomes das sheets que pertencem a planilha 
wb.sheetnames


#interando sobre as planilha 
for w in wb.sheetnames:
    print(w)
    
    
ws.sheet_properties.tabColor="1072BA"

wb.save("Teste.xlsx")


#lendo a planiilha 

wb= load_workbook('Teste.xlsx')

wb.active

#mudando a planilha ativa (que esta na posição zero)
wb.active=wb['Minha planilha']

wb.active



wb=load_workbook('Estudo_ponto.xlsx')

print(wb)

wb.active

teste=wb['teste']