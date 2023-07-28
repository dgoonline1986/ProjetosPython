
#pip install tabula
#pip install Ipython

#ver o pdf no display

import tabula
import pandas as pd
from IPython.display import display
from openpyxl import load_workbook

lista_tabelas = tabula.read_pdf('formatado.pdf', pages='all')

arquivo_excel = 'Estudo_ponto.xlsx'
book = load_workbook(arquivo_excel)

nome_sheet = 'Planilha1'

planilha = book[nome_sheet]
proximo_indice =0
proximo_indice = planilha.max_row + 1


for linha in lista_tabelas:
    df=pd.DataFrame(linha)
    
    valores=df.values.tolist()
    
    for l in valores:
        planilha.append(l)

book.save(arquivo_excel)
    