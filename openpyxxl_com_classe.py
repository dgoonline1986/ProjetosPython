from openpyxl import Workbook
from openpyxl import load_workbook


#criando a classe

class Planilha():
    def __init__(self,wb):
      self.wb=wb
      
      
    def save_wb(self,filename):
        self.wb.save(f'{filename}')
        
    def open_wb(self,filename):
        self.wb=load_workbook(filename)
        
        #return load_workbook(filename)

    def create_sheets(self,sheets):
        for s in sheets:
            self.wb.create_sheet(s)
            
    def delete_sheet_by_name(self,sheetName):
        self.wb.remove(self.wb[sheetName])
        
    def copy_sheet(self,sheetName,new_sheet=''):
        if new_sheet !="":
            ws=self.wb[sheetName]
            n_sheet=self.wb.copy_worksheet(ws)
            n_sheet.title=new_sheet
            
    def get_sheet_by_index(self,index):
        
        try:
            print(self.wb.worksheets[index])
            return self.wb.worksheets[index]
        except IndexError:
            print("planilha não existente",index)
            
    def fill_Data(self,sheet,lista):
        ws=self.wb[sheet]
        for x, data in enumerate(lista):
            for y, item in enumerate(data):
             mw.cell(row=x+1,column=y+1,value=item)
             
             
             
             
#estanciando metodo 

my_plan=Planilha(Workbook())

#criando sheet
my_plan.create_sheets(['teste1','teste2','teste3'])
#salvando
my_plan.save_wb('plan_teste.xlsx')

#escolhendo index paa mostrar 
my_plan.get_sheet_by_index(1)

#deletando
my_plan.delete_sheet_by_name('Sheet')

#salvando
my_plan.save_wb('plan_teste.xlsx')


lista=[
    ['CODIGO','DESCRICAO','VALOR'],
       ['001','MATERIALDE LIMPEZA','1000'],
       ['002','MATERIAL DE CONSUMO','2000'],
       ['003','MATERIA PRIMA','5000'],
       
       ]

#inserindo a lista na planilha passando nome da sheet e os dados 

my_plan.fill_Data('teste1',lista)

#salvando
my_plan.save_wb('plan_teste.xlsx')




# proximo video https://www.youtube.com/watch?v=SkA5nC11zso
